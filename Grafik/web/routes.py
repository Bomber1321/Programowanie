from flask import Blueprint, render_template, g, redirect, url_for, flash, request, send_file, jsonify, abort
from db.init import get_db 
import sqlite3
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

web = Blueprint("web", __name__)


@web.route("/")
def index():
    return render_template("index.html")

@web.route("/ping-db")
def ping_db():
    db = get_db()
    db.execute("SELECT 1").fetchone()
    return render_template("ping.html")

@web.route("/add_month", methods=["GET", "POST"])
def add_month():
    db = get_db()
    cursor = db.cursor()

    if request.method == "POST":
        month_name = request.form.get("month")
        month_length = int(request.form.get("month_length"))
        first_day = request.form.get("first_day")
        year = int(request.form.get("year"))

        cursor.execute("INSERT INTO months(name, length, weekday_start, year) VALUES (?, ?, ?, ?)", 
                       (month_name, month_length, first_day, year))
        

        new_month_id = cursor.lastrowid 

        weekdays = ['Poniedziałek', 'Wtorek', 'Środa', 'Czwartek', 'Piątek', 'Sobota', 'Niedziela']
        
        try:
            start_index = weekdays.index(first_day.capitalize())
        except ValueError:
            start_index = 0
            

        days_data = []
        for day_index in range(1, month_length + 1):
            current_weekday = weekdays[(start_index + day_index - 1) % 7]
            days_data.append((new_month_id, day_index, current_weekday, 1))
        

        cursor.executemany("""
            INSERT INTO calendar (month_id, day_index, weekday, available) 
            VALUES (?, ?, ?, ?)
        """, days_data)
        

        employees = db.execute("SELECT id FROM employees").fetchall()
        availability_data = []
        for emp in employees:
            emp_id = emp['id']
            for day_index in range(1, month_length + 1):
                current_weekday = weekdays[(start_index + day_index - 1) % 7]
                availability_data.append((emp_id, new_month_id, day_index, current_weekday, 0))  # Default to unavailable
        

        cursor.executemany("""
            INSERT INTO availability (id_employee, id_month, day, weekday, available) 
            VALUES (?, ?, ?, ?, ?)
        """, availability_data)

        db.commit()
        return redirect(url_for("web.add_month"))

    
    if request.method == "GET":
        months = db.execute("SELECT * FROM months").fetchall()
    return render_template("add_month.html", months=months)

@web.route("/list_employees", methods=["GET","POST"])
def list_employees():
    db = get_db()
    positions = db.execute("SELECT name FROM positions").fetchall()
    if request.method == "POST":
        name = request.form.get("name")
        surname = request.form.get("surname")
        main_position = request.form.get("main_position")
  
        
        if main_position not in [position['name'] for position in positions]:
            flash("Nieprawidłowe stanowisko.", "danger")
            return redirect(url_for("web.list_employees"))
        
        
        db.execute("INSERT INTO employees (name, surname, positions) VALUES (?, ?, ?)", 
                   (name, surname, main_position))
        db.commit()
        flash("Pracownik został dodany.", "success")
        return redirect(url_for("web.list_employees"))
    raw_employees = db.execute("SELECT id, name, surname, positions, active, created_at FROM employees ORDER by created_at DESC").fetchall()
        
    return render_template("list_employees.html", employees=raw_employees, positions=positions)

@web.route("/positions", methods=["GET","POST"])
def positions():
    db = get_db()
    positions = db.execute("SELECT * FROM positions").fetchall()
    if request.method == "POST":
        position_name = request.form.get("position_name")

        if position_name:
            if position_name in [position['name'] for position in positions]:
                flash("Stanowisko o tej nazwie już istnieje.", "danger")
            else:
                db.execute("INSERT INTO positions (name) VALUES (?)", (position_name,))
                db.commit()
                positions = db.execute("SELECT * FROM positions").fetchall()
                for position in positions:
                    if position['name'] == position_name:
                        position_id = position['id']
                        db.executemany("INSERT INTO required_employees (weekday, position, required_count) VALUES (?, ?, ?)",
                                       [["Poniedziałek", position_id, 0],
                                        ["Wtorek", position_id, 0],
                                        ["Środa", position_id, 0],
                                        ["Czwartek", position_id, 0],
                                        ["Piątek", position_id, 0],
                                        ["Sobota", position_id, 0],
                                        ["Niedziela", position_id, 0]])
                        db.commit()
                        flash("Stanowisko zostało dodane, uzupełnij wymagania!", "success")
    positions = db.execute("SELECT * FROM positions").fetchall()
    return render_template("positions.html", positions=positions)

@web.route("/delete_position/<int:position_id>", methods=["POST"])
def delete_position(position_id):
    db = get_db()
    position = db.execute("SELECT name FROM positions WHERE id = ?", (position_id,)).fetchone()
    if position is None:
        flash("Nie znaleziono stanowiska o podanym ID.", "danger")
        return redirect(url_for("web.positions"))
    db.execute("DELETE FROM required_employees WHERE position = ?", (position_id,))
    db.execute("DELETE FROM positions WHERE id = ?", (position_id,))
    db.commit()
    flash(f"Stanowisko '{position['name']}' zostało usunięte.", "success")
    return redirect(url_for("web.positions"))

@web.route("/edit_employee/<int:employee_id>", methods=["GET","POST"])
def edit_employee(employee_id):
    db = get_db()    
    positions = db.execute("SELECT * FROM positions").fetchall()

    employee = db.execute("SELECT id, name, surname, positions, active, created_at FROM employees WHERE id = ?", (employee_id,)).fetchone()
    if employee is None:
        flash("Nie znaleziono pracownika o podanym ID.", "danger")
        return redirect(url_for("web.list_employees"))
    
    
    return render_template("edit_employee.html", employee=employee, positions=positions)

@web.route("/edit_employee/<int:employee_id>/main_position", methods=["POST"])
def edit_employee_main_position(employee_id):
    db = get_db()
    new_main_position = request.form.get("main_position")
    positions = db.execute("SELECT * FROM positions").fetchall()
    if new_main_position not in [position['name'] for position in positions]:
        flash("Nieprawidłowa pozycja główna.", "danger")
        return redirect(url_for("web.edit_employee", employee_id=employee_id))
    
    employee = db.execute("SELECT positions FROM employees WHERE id = ?", (employee_id,)).fetchone()
    if employee is None:
        flash("Nie znaleziono pracownika o podanym ID.", "danger")
        return redirect(url_for("web.list_employees"))
    
    
    db.execute("UPDATE employees SET positions = ? WHERE id = ?", (new_main_position, employee_id))
    db.commit()
    
    flash("Główna pozycja została zaktualizowana.", "success")
    return redirect(url_for("web.edit_employee", employee_id=employee_id))

@web.route("/edit_employee/<int:employee_id>/active", methods=["POST"])
def edit_employee_active(employee_id):
    db = get_db()
    new_active_status = request.form.get("active")
    
    if new_active_status not in ['0', '1']:
        flash("Nieprawidłowy status aktywności.", "danger")
        return redirect(url_for("web.edit_employee", employee_id=employee_id))
    
    db.execute("UPDATE employees SET active = ? WHERE id = ?", (int(new_active_status), employee_id))
    db.commit()
    
    flash("Status aktywności został zaktualizowany.", "success")
    return redirect(url_for("web.edit_employee", employee_id=employee_id))

@web.route("/delete_employee/<int:employee_id>", methods=["POST"])
def delete_employee(employee_id):
    db = get_db()
    db.execute("DELETE FROM availability WHERE id_employee = ?", (employee_id,))
    db.execute("DELETE FROM employees WHERE id = ?", (employee_id,))
    db.commit()
    flash("Pracownik został usunięty.", "success")
    return redirect(url_for("web.list_employees"))

@web.route("/month/<int:month_id>/availability", methods=["GET", "POST"])
def month_availability(month_id):
    db = get_db()
    month_calendar = db.execute("SELECT * FROM calendar WHERE month_id = ?", (month_id,)).fetchone()
    month = db.execute("SELECT * FROM months WHERE id = ?", (month_id,)).fetchone()
    
    if month is None:
        flash("Nie znaleziono miesiąca o podanym ID.", "danger")
        return redirect(url_for("web.add_month"))
        

    calendar = db.execute("SELECT * FROM calendar WHERE month_id = ? ORDER BY day_index", (month_id,)).fetchall()
    employees = db.execute("SELECT id, active, name, positions, surname FROM employees").fetchall()
    positions = db.execute("SELECT * FROM positions").fetchall()
    positions_length = db.execute("SELECT COUNT(*) FROM positions").fetchone()[0]
    requirements = db.execute("SELECT * FROM required_employees").fetchall()

    availability = db.execute("SELECT * FROM availability WHERE id_month = ? ORDER BY id", (month_id,)).fetchall()
    existing_keys = {(row['id_employee'], row['day']) for row in availability}
    missing_availability = []
    for employee in employees:
        for day in calendar:
            if (employee['id'], day['day_index']) not in existing_keys:
                missing_availability.append((employee['id'], month_id, day['day_index'], day['weekday'], 0))

    if missing_availability:
        db.executemany(
            "INSERT INTO availability (id_employee, id_month, day, weekday, available) VALUES (?, ?, ?, ?, ?)",
            missing_availability
        )
        db.commit()
        availability = db.execute("SELECT * FROM availability WHERE id_month = ? ORDER BY id", (month_id,)).fetchall()

    positions_by_name = {position['name']: position['id'] for position in positions}
    enriched_employees = []
    for employee in employees:
        position_name = employee['positions']
        position_id = positions_by_name.get(position_name, 0)
        enriched_employee = dict(employee)
        enriched_employee['position_id'] = position_id
        enriched_employees.append(enriched_employee)

    availability_map = {f"{row['id_employee']}_{row['day']}": row['available'] for row in availability}

    requirements_by_weekday_position = {}
    for requirement in requirements:
        requirements_by_weekday_position.setdefault(requirement['weekday'], {})[requirement['position']] = requirement['required_count']

    return render_template(
        "month_availability.html",
        month_info=month,
        calendar=calendar,
        employees=enriched_employees,
        availability_map=availability_map,
        requirements=requirements,
        requirements_by_weekday_position=requirements_by_weekday_position,
        positions=positions,
        positions_length=positions_length,
    )

@web.route("/month/<int:month_id>/edit_availability", methods=["POST"])
def edit_availability(month_id):
    db = get_db()
    month = db.execute("SELECT * FROM months WHERE id = ?", (month_id,)).fetchone()
    
    if month is None:
        flash("Nie znaleziono miesiąca o podanym ID.", "danger")
        return redirect(url_for("web.add_month"))
    
    employees = db.execute("SELECT id FROM employees").fetchall()
    calendar = db.execute("SELECT day_index, weekday FROM calendar WHERE month_id = ? ORDER BY day_index", (month_id,)).fetchall()
    existing_availability = {(row['id_employee'], row['day']) for row in db.execute("SELECT id_employee, day FROM availability WHERE id_month = ?", (month_id,)).fetchall()}
    missing_availability = []
    for employee in employees:
        for day in calendar:
            if (employee['id'], day['day_index']) not in existing_availability:
                missing_availability.append((employee['id'], month_id, day['day_index'], day['weekday'], 0))

    if missing_availability:
        db.executemany(
            "INSERT INTO availability (id_employee, id_month, day, weekday, available) VALUES (?, ?, ?, ?, ?)",
            missing_availability
        )

    db.execute("UPDATE availability SET available = 0 WHERE id_month = ?", (month_id,))
    
    for key, value in request.form.items():
        if key.startswith("availability_"):
            _, emp_id, day_index = key.split("_")
            emp_id = int(emp_id)
            day_index = int(day_index)
            available = 1 if value == "1" else 0
            
            db.execute("""
                UPDATE availability 
                SET available = ? 
                WHERE id_employee = ? AND id_month = ? AND day = ?
            """, (available, emp_id, month_id, day_index))
    
    db.commit()
    flash("Dostępność została zaktualizowana.", "success")
    return redirect(url_for("web.month_availability", month_id=month_id))

@web.route("/download-availability/<int:month_id>", methods=["GET"])
def download_availability(month_id):
    db = get_db()
    availability_schedule = db.execute("SELECT availability.id, availability.id_month, availability.day, availability.weekday, concat(employees.name,' ',employees.surname) Fullname, availability.available FROM availability INNER JOIN employees on employees.id = availability.id_employee WHERE availability.id_month = ?", (month_id,)).fetchall()
    availability_schedule_unique_names = list(set(record['Fullname'] for record in availability_schedule))
    month_length = db.execute("SELECT length FROM months WHERE id = ?", (month_id,)).fetchone()[0]
    month_name = db.execute("SELECT name FROM months WHERE id = ?", (month_id,)).fetchone()[0]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month_name

    headers = ["Data", "Dzień Tygodnia"] + availability_schedule_unique_names
    ws.append(headers)


    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for day in range(0, month_length):
        row = [
            day + 1,
            availability_schedule[day][3]
        ]
        for name in availability_schedule_unique_names:
            record = next((rec for rec in availability_schedule if rec['Fullname'] == name and rec['day'] == day + 1), None)
            if record:
                row.append("O" if record['available'] == 1 else "")
            else:
                row.append("Brak danych")
        ws.append(row) 
        
    uniform_width = 20 

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx) 
        ws.column_dimensions[col_letter].width = uniform_width
        
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14


    memory_file = io.BytesIO()
    wb.save(memory_file)
    
    memory_file.seek(0)

    return send_file(
        memory_file,
        as_attachment=True,
        download_name='employee_availability.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@web.route("/requirements")
def requirements():
    db = get_db()
    requirements = db.execute("""
        SELECT r.id, r.weekday, p.name AS position_name, r.required_count
        FROM required_employees r
        JOIN positions p ON r.position = p.id
        ORDER BY p.name
    """).fetchall()
    
    return render_template("required_employees.html", requirements=requirements)

@web.route("/edit_requirement/<int:requirement_id>", methods=["POST"])
def edit_requirement(requirement_id):
    db = get_db()
    requirement = db.execute("SELECT * FROM required_employees WHERE id = ?", (requirement_id,)).fetchone()

    if requirement is None:
        flash("Nie znaleziono wymagania o podanym ID.", "danger")
        return redirect(url_for("web.requirements"))

    new_count = request.form.get("required_count")
    if new_count is None:
        flash("Nieprawidłowa wartość dla wymaganej liczby pracowników.", "danger")
        return redirect(url_for("web.requirements"))

    try:
        new_count = int(new_count)
        if new_count < 0:
            raise ValueError("Liczba pracowników nie może być ujemna.")
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("web.requirements"))

    db.execute("UPDATE required_employees SET required_count = ? WHERE id = ?", (new_count, requirement_id))
    db.commit()

    flash("Wymaganie zostało zaktualizowane.", "success")
    return redirect(url_for("web.requirements"))
