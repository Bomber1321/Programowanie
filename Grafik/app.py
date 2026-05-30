from flask import Flask, render_template, g, redirect, url_for, flash, request, send_file, jsonify, abort
import secrets
import sqlite3
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.config['SECRET_KEY'] = secrets.token_urlsafe(16)
DATABASE = "todo.db"
SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS employees(
id INTEGER PRIMARY KEY AUTOINCREMENT,
name TEXT NOT NULL,
surname TEXT NOT NULL,
positions TEXT not null,
active INTEGER NOT NULL DEFAULT 1 CHECK (active IN (0, 1)),
created_at TEXT NOT NULL DEFAULT (datetime('now'))
);

CREATE INDEX IF NOT EXISTS idx_employees_active ON employees(active);
CREATE INDEX IF NOT EXISTS idx_employees_created_at ON employees(created_at);


CREATE TABLE IF NOT EXISTS availability(
id INTEGER PRIMARY KEY AUTOINCREMENT,
id_employee INTEGER NOT NULL,
id_month INTEGER NOT NULL,
day INTEGER NOT NULL,
weekday TEXT NOT NULL,
available INTEGER NOT NULL DEFAULT 0 CHECK (available IN (0, 1)),
FOREIGN KEY(id_employee) REFERENCES employees(id),
FOREIGN KEY(id_month) REFERENCES months(id)
);

CREATE TABLE IF NOT EXISTS months (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL,
    length INTEGER NOT NULL,
    weekday_start TEXT NOT NULL,
    year INTEGER NOT NULL
);

CREATE TABLE IF NOT EXISTS calendar (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    month_id INTEGER NOT NULL,
    day_index INTEGER NOT NULL,
    weekday TEXT NOT NULL,
    available INTEGER DEFAULT 0 CHECK (available IN (0, 1)),
    FOREIGN KEY (month_id) REFERENCES months (id)
);

CREATE TABLE IF NOT EXISTS positions (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL
);

create TABLE IF NOT EXISTS required_employees (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    weekday TEXT NOT NULL,
    position INTEGER NOT NULL,
    required_count INTEGER NOT NULL DEFAULT 0,
    FOREIGN KEY (position) REFERENCES positions(id)
    )
"""
def SEED_SQL():
    db = get_db()
    seed_employees = db.executemany("INSERT INTO employees(name, surname, positions, active) VALUES (?, ?, ?, ?)",
    [["Norbert", "Czerw", 'Kelner', 1],
    ["Joanna", "Czerw", 'Barman', 1],
    ["Mateusz", "Cieślik", 'Pizzer', 1],
    ["Daria", "Rogala", 'Pizzer', 1],
    ["Paulina", "Smusz", 'Pizzer', 1],
    ["Wiktoria", "Kisielewska", 'Barman', 1],
    ["Zuzanna", "Adamska", 'Kelner', 1],
    ["Jakub", "Kura", 'Kierowca', 1]])
    
    seed_positions = db.executemany("INSERT INTO positions(name) VALUES (?)",
    [["Pizzer"], ["Barman"], ["Kelner"], ["Kierowca"]])
    
    seed_required_employees = db.executemany("INSERT INTO required_employees(weekday, position, required_count) VALUES (?, ?, ?)",
    [["Poniedziałek", 1, 2],
    ["Poniedziałek", 2, 1],
    ["Poniedziałek", 3, 2],
    ["Poniedziałek", 4, 1],
    ["Wtorek", 1, 2],
    ["Wtorek", 2, 1],
    ["Wtorek", 3, 2],
    ["Wtorek", 4, 1],
    ["Środa", 1, 2],
    ["Środa", 2, 1],
    ["Środa", 3, 2],
    ["Środa", 4, 1],
    ["Czwartek", 1, 2],
    ["Czwartek", 2, 1],
    ["Czwartek", 3, 2],
    ["Czwartek", 4, 1],
    ["Piątek", 1, 3],
    ["Piątek", 2, 2],
    ["Piątek", 3, 3],
    ["Piątek", 4, 2],
    ["Sobota", 1, 3],
    ["Sobota", 2, 2],
    ["Sobota", 3, 3],
    ["Sobota", 4, 2],
    ["Niedziela", 1, 3],
    ["Niedziela", 2, 2],
    ["Niedziela", 3, 3],
    ["Niedziela", 4, 2]])
    
    return 0


def get_db():
    if "db" not in g:
        conn = sqlite3.connect(DATABASE)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON;")
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(exception):
    db = g.pop("db", None)
    if db is not None:
        db.close()

def init_db():
    db = get_db()
    db.executescript(SCHEMA_SQL)
    db.commit()

@app.cli.command("init-db")
def init_db_command():
    init_db()
    print("✔ Zainicjowano bazę danych")
    
@app.cli.command("seed-db")
def seed_db_command():
    db = get_db()
    isDatabaseEmpty = db.execute("SELECT COUNT(*) FROM employees").fetchone()[0] + db.execute("SELECT COUNT(*) FROM positions").fetchone()[0] + db.execute("SELECT COUNT(*) FROM required_employees").fetchone()[0]
    if isDatabaseEmpty == 0:
        SEED_SQL()
        db.commit()
        print("✔ Dane przykładowe zostały dodane do tabeli employees.")
    else:
        print("❌ Tabela employees zawiera już dane, seedowanie przerwane.")
        
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/ping-db")
def ping_db():
    db = get_db()
    db.execute("SELECT 1").fetchone()
    return render_template("ping.html")

@app.route("/add_month", methods=["GET", "POST"])
def add_month():
    db = get_db()
    cursor = db.cursor()

    if request.method == "POST":
        month_name = request.form.get("month")
        month_length = int(request.form.get("month_length"))
        first_day = request.form.get("first_day")
        year = int(request.form.get("year"))

        cursor.execute("INSERT INTO months(name, length, weekday_start, year) VALUES (?, ?, ?, ?)", 
                       (month_name, month_length, first_day, year))
        

        new_month_id = cursor.lastrowid 

        weekdays = ['Poniedziałek', 'Wtorek', 'Środa', 'Czwartek', 'Piątek', 'Sobota', 'Niedziela']
        
        try:
            start_index = weekdays.index(first_day.capitalize())
        except ValueError:
            start_index = 0
            

        days_data = []
        for day_index in range(1, month_length + 1):
            current_weekday = weekdays[(start_index + day_index - 1) % 7]
            days_data.append((new_month_id, day_index, current_weekday, 1))
        

        cursor.executemany("""
            INSERT INTO calendar (month_id, day_index, weekday, available) 
            VALUES (?, ?, ?, ?)
        """, days_data)
        

        employees = db.execute("SELECT id FROM employees").fetchall()
        availability_data = []
        for emp in employees:
            emp_id = emp['id']
            for day_index in range(1, month_length + 1):
                current_weekday = weekdays[(start_index + day_index - 1) % 7]
                availability_data.append((emp_id, new_month_id, day_index, current_weekday, 0))  # Default to unavailable
        

        cursor.executemany("""
            INSERT INTO availability (id_employee, id_month, day, weekday, available) 
            VALUES (?, ?, ?, ?, ?)
        """, availability_data)

        db.commit()
        return redirect(url_for("add_month"))

    
    if request.method == "GET":
        months = db.execute("SELECT * FROM months").fetchall()
    return render_template("add_month.html", months=months)

@app.route("/list_employees", methods=["GET","POST"])
def list_employees():
    db = get_db()
    positions = db.execute("SELECT name FROM positions").fetchall()
    if request.method == "POST":
        name = request.form.get("name")
        surname = request.form.get("surname")
        main_position = request.form.get("main_position")
  
        
        if main_position not in [position['name'] for position in positions]:
            flash("Nieprawidłowe stanowisko.", "danger")
            return redirect(url_for("list_employees"))
        
        
        db.execute("INSERT INTO employees (name, surname, positions) VALUES (?, ?, ?)", 
                   (name, surname, main_position))
        db.commit()
        flash("Pracownik został dodany.", "success")
        return redirect(url_for("list_employees"))
    raw_employees = db.execute("SELECT id, name, surname, positions, active, created_at FROM employees ORDER by created_at DESC").fetchall()
        
    return render_template("list_employees.html", employees=raw_employees, positions=positions)

@app.route("/positions", methods=["GET","POST"])
def positions():
    db = get_db()
    positions = db.execute("SELECT * FROM positions").fetchall()
    if request.method == "POST":
        position_name = request.form.get("position_name")

        if position_name:
            if position_name in [position['name'] for position in positions]:
                flash("Stanowisko o tej nazwie już istnieje.", "danger")
            else:
                db.execute("INSERT INTO positions (name) VALUES (?)", (position_name,))
                db.commit()
                positions = db.execute("SELECT * FROM positions").fetchall()
                for position in positions:
                    if position['name'] == position_name:
                        position_id = position['id']
                        db.executemany("INSERT INTO required_employees (weekday, position, required_count) VALUES (?, ?, ?)",
                                       [["Poniedziałek", position_id, 0],
                                        ["Wtorek", position_id, 0],
                                        ["Środa", position_id, 0],
                                        ["Czwartek", position_id, 0],
                                        ["Piątek", position_id, 0],
                                        ["Sobota", position_id, 0],
                                        ["Niedziela", position_id, 0]])
                        db.commit()
                        flash("Stanowisko zostało dodane, uzupełnij wymagania!", "success")
    positions = db.execute("SELECT * FROM positions").fetchall()
    return render_template("positions.html", positions=positions)

@app.route("/delete_position/<int:position_id>", methods=["POST"])
def delete_position(position_id):
    db = get_db()
    position = db.execute("SELECT name FROM positions WHERE id = ?", (position_id,)).fetchone()
    if position is None:
        flash("Nie znaleziono stanowiska o podanym ID.", "danger")
        return redirect(url_for("positions"))
    db.execute("DELETE FROM required_employees WHERE position = ?", (position_id,))
    db.execute("DELETE FROM positions WHERE id = ?", (position_id,))
    db.commit()
    flash(f"Stanowisko '{position['name']}' zostało usunięte.", "success")
    return redirect(url_for("positions"))

@app.route("/edit_employee/<int:employee_id>", methods=["GET","POST"])
def edit_employee(employee_id):
    db = get_db()    
    positions = db.execute("SELECT * FROM positions").fetchall()

    employee = db.execute("SELECT id, name, surname, positions, active, created_at FROM employees WHERE id = ?", (employee_id,)).fetchone()
    if employee is None:
        flash("Nie znaleziono pracownika o podanym ID.", "danger")
        return redirect(url_for("list_employees"))
    
    
    return render_template("edit_employee.html", employee=employee, positions=positions)

@app.route("/edit_employee/<int:employee_id>/main_position", methods=["POST"])
def edit_employee_main_position(employee_id):
    db = get_db()
    new_main_position = request.form.get("main_position")
    positions = db.execute("SELECT * FROM positions").fetchall()
    if new_main_position not in [position['name'] for position in positions]:
        flash("Nieprawidłowa pozycja główna.", "danger")
        return redirect(url_for("edit_employee", employee_id=employee_id))
    
    employee = db.execute("SELECT positions FROM employees WHERE id = ?", (employee_id,)).fetchone()
    if employee is None:
        flash("Nie znaleziono pracownika o podanym ID.", "danger")
        return redirect(url_for("list_employees"))
    
    
    db.execute("UPDATE employees SET positions = ? WHERE id = ?", (new_main_position, employee_id))
    db.commit()
    
    flash("Główna pozycja została zaktualizowana.", "success")
    return redirect(url_for("edit_employee", employee_id=employee_id))

@app.route("/edit_employee/<int:employee_id>/active", methods=["POST"])
def edit_employee_active(employee_id):
    db = get_db()
    new_active_status = request.form.get("active")
    
    if new_active_status not in ['0', '1']:
        flash("Nieprawidłowy status aktywności.", "danger")
        return redirect(url_for("edit_employee", employee_id=employee_id))
    
    db.execute("UPDATE employees SET active = ? WHERE id = ?", (int(new_active_status), employee_id))
    db.commit()
    
    flash("Status aktywności został zaktualizowany.", "success")
    return redirect(url_for("edit_employee", employee_id=employee_id))

@app.route("/delete_employee/<int:employee_id>", methods=["POST"])
def delete_employee(employee_id):
    db = get_db()
    db.execute("DELETE FROM employees WHERE id = ?", (employee_id,))
    db.commit()
    flash("Pracownik został usunięty.", "success")
    return redirect(url_for("list_employees"))

@app.route("/month/<int:month_id>/availability", methods=["GET", "POST"])
def month_availability(month_id):
    db = get_db()
    month_calendar = db.execute("SELECT * FROM calendar WHERE month_id = ?", (month_id,)).fetchone()
    month = db.execute("SELECT * FROM months WHERE id = ?", (month_id,)).fetchone()
    
    if month is None:
        flash("Nie znaleziono miesiąca o podanym ID.", "danger")
        return redirect(url_for("add_month"))
        

    calendar = db.execute("SELECT * FROM calendar WHERE month_id = ? ORDER BY day_index", (month_id,)).fetchall()
    employees = db.execute("SELECT id, active, name, positions, surname FROM employees").fetchall()
    positions = db.execute("SELECT * FROM positions").fetchall()
    positions_length = db.execute("SELECT COUNT(*) FROM positions").fetchone()[0]
    requirements = db.execute("SELECT * FROM required_employees").fetchall()

    availability = db.execute("SELECT * FROM availability WHERE id_month = ? ORDER BY id", (month_id,)).fetchall()
    existing_keys = {(row['id_employee'], row['day']) for row in availability}
    missing_availability = []
    for employee in employees:
        for day in calendar:
            if (employee['id'], day['day_index']) not in existing_keys:
                missing_availability.append((employee['id'], month_id, day['day_index'], day['weekday'], 0))

    if missing_availability:
        db.executemany(
            "INSERT INTO availability (id_employee, id_month, day, weekday, available) VALUES (?, ?, ?, ?, ?)",
            missing_availability
        )
        db.commit()
        availability = db.execute("SELECT * FROM availability WHERE id_month = ? ORDER BY id", (month_id,)).fetchall()

    positions_by_name = {position['name']: position['id'] for position in positions}
    enriched_employees = []
    for employee in employees:
        position_name = employee['positions']
        position_id = positions_by_name.get(position_name, 0)
        enriched_employee = dict(employee)
        enriched_employee['position_id'] = position_id
        enriched_employees.append(enriched_employee)

    availability_map = {f"{row['id_employee']}_{row['day']}": row['available'] for row in availability}

    requirements_by_weekday_position = {}
    for requirement in requirements:
        requirements_by_weekday_position.setdefault(requirement['weekday'], {})[requirement['position']] = requirement['required_count']

    return render_template(
        "month_availability.html",
        month_info=month,
        calendar=calendar,
        employees=enriched_employees,
        availability_map=availability_map,
        requirements=requirements,
        requirements_by_weekday_position=requirements_by_weekday_position,
        positions=positions,
        positions_length=positions_length,
    )

@app.route("/month/<int:month_id>/edit_availability", methods=["POST"])
def edit_availability(month_id):
    db = get_db()
    month = db.execute("SELECT * FROM months WHERE id = ?", (month_id,)).fetchone()
    
    if month is None:
        flash("Nie znaleziono miesiąca o podanym ID.", "danger")
        return redirect(url_for("add_month"))
    
    employees = db.execute("SELECT id FROM employees").fetchall()
    calendar = db.execute("SELECT day_index, weekday FROM calendar WHERE month_id = ? ORDER BY day_index", (month_id,)).fetchall()
    existing_availability = {(row['id_employee'], row['day']) for row in db.execute("SELECT id_employee, day FROM availability WHERE id_month = ?", (month_id,)).fetchall()}
    missing_availability = []
    for employee in employees:
        for day in calendar:
            if (employee['id'], day['day_index']) not in existing_availability:
                missing_availability.append((employee['id'], month_id, day['day_index'], day['weekday'], 0))

    if missing_availability:
        db.executemany(
            "INSERT INTO availability (id_employee, id_month, day, weekday, available) VALUES (?, ?, ?, ?, ?)",
            missing_availability
        )

    db.execute("UPDATE availability SET available = 0 WHERE id_month = ?", (month_id,))
    
    for key, value in request.form.items():
        if key.startswith("availability_"):
            _, emp_id, day_index = key.split("_")
            emp_id = int(emp_id)
            day_index = int(day_index)
            available = 1 if value == "1" else 0
            
            db.execute("""
                UPDATE availability 
                SET available = ? 
                WHERE id_employee = ? AND id_month = ? AND day = ?
            """, (available, emp_id, month_id, day_index))
    
    db.commit()
    flash("Dostępność została zaktualizowana.", "success")
    return redirect(url_for("month_availability", month_id=month_id))

@app.route("/download-availability/<int:month_id>", methods=["GET"])
def download_availability(month_id):
    db = get_db()
    availability_schedule = db.execute("SELECT availability.id, availability.id_month, availability.day, availability.weekday, concat(employees.name,' ',employees.surname) Fullname, availability.available FROM availability INNER JOIN employees on employees.id = availability.id_employee WHERE availability.id_month = ?", (month_id,)).fetchall()
    availability_schedule_unique_names = list(set(record['Fullname'] for record in availability_schedule))
    month_length = db.execute("SELECT length FROM months WHERE id = ?", (month_id,)).fetchone()[0]
    month_name = db.execute("SELECT name FROM months WHERE id = ?", (month_id,)).fetchone()[0]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month_name

    headers = ["Data", "Dzień Tygodnia"] + availability_schedule_unique_names
    ws.append(headers)


    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    for day in range(0, month_length):
        row = [
            day + 1,
            availability_schedule[day][3]
        ]
        for name in availability_schedule_unique_names:
            record = next((rec for rec in availability_schedule if rec['Fullname'] == name and rec['day'] == day + 1), None)
            if record:
                row.append("O" if record['available'] == 1 else "")
            else:
                row.append("Brak danych")
        ws.append(row) 
        
    uniform_width = 20 

    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx) 
        ws.column_dimensions[col_letter].width = uniform_width
        
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 14


    memory_file = io.BytesIO()
    wb.save(memory_file)
    
    memory_file.seek(0)

    return send_file(
        memory_file,
        as_attachment=True,
        download_name='employee_availability.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route("/requirements")
def requirements():
    db = get_db()
    requirements = db.execute("""
        SELECT r.id, r.weekday, p.name AS position_name, r.required_count
        FROM required_employees r
        JOIN positions p ON r.position = p.id
        ORDER BY p.name
    """).fetchall()
    
    return render_template("required_employees.html", requirements=requirements)

@app.route("/edit_requirement/<int:requirement_id>", methods=["POST"])
def edit_requirement(requirement_id):
    db = get_db()
    requirement = db.execute("SELECT * FROM required_employees WHERE id = ?", (requirement_id,)).fetchone()

    if requirement is None:
        flash("Nie znaleziono wymagania o podanym ID.", "danger")
        return redirect(url_for("requirements"))

    new_count = request.form.get("required_count")
    if new_count is None:
        flash("Nieprawidłowa wartość dla wymaganej liczby pracowników.", "danger")
        return redirect(url_for("requirements"))

    try:
        new_count = int(new_count)
        if new_count < 0:
            raise ValueError("Liczba pracowników nie może być ujemna.")
    except ValueError as e:
        flash(str(e), "danger")
        return redirect(url_for("requirements"))

    db.execute("UPDATE required_employees SET required_count = ? WHERE id = ?", (new_count, requirement_id))
    db.commit()

    flash("Wymaganie zostało zaktualizowane.", "success")
    return redirect(url_for("requirements"))



#--------------------------------------------------------------------------------------------------------------------------------------#
#-------------------------------------------------------------------API----------------------------------------------------------------#
#--------------------------------------------------------------------------------------------------------------------------------------#



@app.route("/employees/api", methods=["GET"])
def api_list_employees():
    db = get_db()
    rows = db.execute("SELECT id, name, surname, positions, active, created_at FROM employees ORDER by created_at DESC").fetchall()
    return jsonify([dict(row) for row in rows])

@app.route("/employees/api/<int:employee_id>", methods=["GET"])
def api_employees_get(employee_id):
    db = get_db()
    row = db.execute("SELECT id, name, surname, positions, active, created_at FROM employees WHERE id = ?",[employee_id]).fetchone()

    if row is None:
        abort(404, description="Employee not found")
    
    return jsonify(dict(row))

@app.route("/employees/api", methods=["POST"])
def api_employee_add():
    db = get_db()
    data = request.get_json(silent=True)

    if not data or "name" not in data or "surname" not in data or "positions" not in data or "active" not in data:
        abort(400, description="Missing DATA")
        
    active = 1 if data["active"] == "active" else 0
    row = db.execute("INSERT INTO employees (name, surname, positions, active) VALUES (?, ?, ?, ?)",(data["name"],data["surname"],data["positions"], active))
    db.commit()
    return jsonify({
        "message": "Employee added successfully", 
        "id": row.lastrowid
    }), 201
    
@app.route("/employees/api/<int:employee_id>", methods=["PUT", "PATCH"])
def api_employee_update(employee_id):
    db = get_db()
    data = request.get_json(silent=True)

    if not data:
        abort(400, description="Missing DATA")

    # Check if employee exists
    employee = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
    if employee is None:
        abort(404, description="Employee not found")

    # Update employee details
    name = data.get("name", employee["name"])
    surname = data.get("surname", employee["surname"])
    positions = data.get("positions", employee["positions"])
    active = 1 if data.get("active") == "active" else 0

    db.execute("UPDATE employees SET name = ?, surname = ?, positions = ?, active = ? WHERE id = ?", (name, surname, positions, active, employee_id))
    db.commit()

    return jsonify({"message": "Employee updated successfully"})

@app.route("/employees/api/<int:employee_id>", methods=["DELETE"])
def api_employee_delete(employee_id):
    db = get_db()
    employee = db.execute("SELECT * FROM employees WHERE id = ?", (employee_id,)).fetchone()
    if employee is None:
        abort(404, description="Employee not found")

    db.execute("DELETE FROM employees WHERE id = ?", (employee_id,))
    db.commit()

    return jsonify({"message": "Employee deleted successfully"})

@app.route("/months/api", methods=["GET"])
def api_list_months():
    db = get_db()
    rows = db.execute("SELECT * FROM months").fetchall()
    return jsonify([dict(row) for row in rows])

@app.route("/months/api/<int:month_id>", methods=["GET"])
def api_months_get(month_id):
    db = get_db()
    row = db.execute("SELECT * FROM months WHERE id = ?",[month_id]).fetchone()

    if row is None:
        abort(404, description="Month not found")
    
    return jsonify(dict(row))

@app.route("/months/api", methods=["POST"])
def api_month_add():
    db = get_db()
    data = request.get_json(silent=True)

    if not data or "name" not in data or "length" not in data or "year" not in data or "weekday_start" not in data:
        abort(400, description="Missing DATA")
        
    if data["weekday_start"] not in ['Poniedziałek', 'Wtorek', 'Środa', 'Czwartek', 'Piątek', 'Sobota', 'Niedziela']:
        abort(400, description="Incorrect weekday_start")
        
    row = db.execute("INSERT INTO months (name, length, year, weekday_start) VALUES (?, ?, ?, ?)",(data["name"],data["length"],data["year"], data["weekday_start"]))
    db.commit()
    return jsonify({
        "message": "Month added successfully", 
        "id": row.lastrowid
    }), 201
    
@app.route("/months/api/<int:month_id>", methods=["PUT", "PATCH"])
def api_month_update(month_id):
    db = get_db()
    data = request.get_json(silent=True)

    if not data:
        abort(400, description="Missing DATA")

    month = db.execute("SELECT * FROM months WHERE id = ?", (month_id,)).fetchone()
    if month is None:
        abort(404, description="Month not found")

    name = data.get("name", month["name"])
    length = data.get("length", month["length"])
    year = data.get("year", month["year"])
    weekday_start = data.get("weekday_start", month["weekday_start"])

    if weekday_start not in ['Poniedziałek', 'Wtorek', 'Środa', 'Czwartek', 'Piątek', 'Sobota', 'Niedziela']:
        abort(400, description="Incorrect weekday_start")

    db.execute("UPDATE months SET name = ?, length = ?, year = ?, weekday_start = ? WHERE id = ?", (name, length, year, weekday_start, month_id))
    db.commit()

    return jsonify({"message": "Month updated successfully"})

@app.route("/months/api/<int:month_id>", methods=["DELETE"])
def api_month_delete(month_id):
    db = get_db()
    month = db.execute("SELECT * FROM months WHERE id = ?", (month_id,)).fetchone()
    if month is None:
        abort(404, description="Month not found")

    db.execute("DELETE FROM months WHERE id = ?", (month_id,))
    db.commit()

    return jsonify({"message": "Month deleted successfully"})

@app.route("/positions/api", methods=["GET"])
def api_list_positions():
    db = get_db()
    rows = db.execute("SELECT * FROM positions").fetchall()
    return jsonify([dict(row) for row in rows])

@app.route("/positions/api/<int:position_id>", methods=["GET"])
def api_positions_get(position_id):
    db = get_db()
    row = db.execute("SELECT * FROM positions WHERE id = ?",[position_id]).fetchone()

    if row is None:
        abort(404, description="Position not found")
    
    return jsonify(dict(row))

@app.route("/positions/api", methods=["POST"])
def api_position_add():
    db = get_db()
    data = request.get_json(silent=True)

    if not data or "name" not in data:
        abort(400, description="Missing DATA")
        
    row = db.execute("INSERT INTO positions (name) VALUES (?)", (data["name"],))
    db.commit()
    return jsonify({
        "message": "Position added successfully", 
        "id": row.lastrowid
    }), 201
    
@app.route("/positions/api/<int:position_id>", methods=["PUT", "PATCH"])
def api_position_update(position_id):
    db = get_db()
    data = request.get_json(silent=True)

    if not data:
        abort(400, description="Missing DATA")

    position = db.execute("SELECT * FROM positions WHERE id = ?", (position_id,)).fetchone()
    if position is None:
        abort(404, description="Position not found")

    name = data.get("name", position["name"])

    db.execute("UPDATE positions SET name = ? WHERE id = ?", (name, position_id))
    db.commit()

    return jsonify({"message": "Position updated successfully"})
    
@app.route("/positions/api/<int:position_id>", methods=["DELETE"])
def api_position_delete(position_id):
    db = get_db()
    position = db.execute("SELECT * FROM positions WHERE id = ?", (position_id,)).fetchone()
    if position is None:
        abort(404, description="Position not found")

    db.execute("DELETE FROM positions WHERE id = ?", (position_id,))
    db.commit()

    return jsonify({"message": "Position deleted successfully"})

@app.route("/required-employees/api", methods=["GET"])
def api_list_required_employees():
    db = get_db()
    rows = db.execute("SELECT * FROM required_employees").fetchall()
    return jsonify([dict(row) for row in rows])

@app.route("/required-employees/api/<int:required_employees_id>", methods=["GET"])
def api_required_employees_get(required_employees_id):
    db = get_db()
    row = db.execute("SELECT * FROM required_employees WHERE id = ?",[required_employees_id]).fetchone()

    if row is None:
        abort(404, description="Required employees row not found")
    
    return jsonify(dict(row))

@app.route("/required-employees/api", methods=["POST"])
def api_required_employees_add():
    db = get_db()
    data = request.get_json(silent=True)

    if not data or "weekday" not in data or "position" not in data or "required_count" not in data:
        abort(400, description="Missing DATA")
        
    row = db.execute("INSERT INTO required_employees (weekday, position, required_count) VALUES (?, ?, ?)",(data["weekday"],data["position"],data["required_count"]))
    db.commit()
    return jsonify({
        "message": "Required employees row added successfully", 
        "id": row.lastrowid
    }), 201
    
@app.route("/required-employees/api/<int:required_employees_id>", methods=["PUT", "PATCH"])
def api_required_employees_update(required_employees_id):
    db = get_db()
    data = request.get_json(silent=True)

    if not data:
        abort(400, description="Missing DATA")

    requirement = db.execute("SELECT * FROM required_employees WHERE id = ?", (required_employees_id,)).fetchone()
    if requirement is None:
        abort(404, description="Required employees row not found")

    weekday = data.get("weekday", requirement["weekday"])
    position = data.get("position", requirement["position"])
    required_count = data.get("required_count", requirement["required_count"])

    db.execute("UPDATE required_employees SET weekday = ?, position = ?, required_count = ? WHERE id = ?", (weekday, position, required_count, required_employees_id))
    db.commit()

    return jsonify({"message": "Required employees row updated successfully"})

@app.route("/required-employees/api/<int:required_employees_id>", methods=["DELETE"])
def api_required_employees_delete(required_employees_id):
    db = get_db()
    requirement = db.execute("SELECT * FROM required_employees WHERE id = ?", (required_employees_id,)).fetchone()
    if requirement is None:
        abort(404, description="Required employees row not found")

    db.execute("DELETE FROM required_employees WHERE id = ?", (required_employees_id,))
    db.commit()

    return jsonify({"message": "Required employees row deleted successfully"})

@app.route("/availability/api", methods=["GET"])
def api_list_availability():
    db = get_db()
    rows = db.execute("SELECT * FROM availability").fetchall()
    return jsonify([dict(row) for row in rows])

@app.route("/availability/api", methods=["POST"])
def api_availability_add():
    db = get_db()
    data = request.get_json(silent=True)

    if not data or "id_employee" not in data or "id_month" not in data or "day" not in data or "weekday" not in data or "available" not in data:
        abort(400, description="Missing DATA")
        
    available = 1 if data["available"] == "available" else 0
    row = db.execute("INSERT INTO availability (id_employee, id_month, day, weekday, available) VALUES (?, ?, ?, ?, ?)",(data["id_employee"],data["id_month"],data["day"],data["weekday"], available))
    db.commit()
    return jsonify({
        "message": "Availability row added successfully", 
        "id": row.lastrowid
    }), 201

@app.route("/availability/api/<int:availability_id>", methods=["GET"])
def api_availability_get(availability_id):
    db = get_db()
    row = db.execute("SELECT * FROM availability WHERE id = ?",[availability_id]).fetchone()

    if row is None:
        abort(404, description="Availability row not found")
    
    return jsonify(dict(row))

@app.route("/availability/api/<int:availability_id>", methods=["PUT", "PATCH"])
def api_availability_update(availability_id):
    db = get_db()
    data = request.get_json(silent=True)

    if not data:
        abort(400, description="Missing DATA")

    availability = db.execute("SELECT * FROM availability WHERE id = ?", (availability_id,)).fetchone()
    if availability is None:
        abort(404, description="Availability row not found")

    id_employee = data.get("id_employee", availability["id_employee"])
    id_month = data.get("id_month", availability["id_month"])
    day = data.get("day", availability["day"])
    weekday = data.get("weekday", availability["weekday"])
    available = 1 if data.get("available") == "available" else 0

    db.execute("UPDATE availability SET id_employee = ?, id_month = ?, day = ?, weekday = ?, available = ? WHERE id = ?", (id_employee, id_month, day, weekday, available, availability_id))
    db.commit()

    return jsonify({"message": "Availability row updated successfully"})

@app.route("/availability/api/<int:availability_id>", methods=["DELETE"])
def api_availability_delete(availability_id):
    db = get_db()
    availability = db.execute("SELECT * FROM availability WHERE id = ?", (availability_id,)).fetchone()
    if availability is None:
        abort(404, description="Availability row not found")

    db.execute("DELETE FROM availability WHERE id = ?", (availability_id,))
    db.commit()

    return jsonify({"message": "Availability row deleted successfully"})

if __name__ == "__main__":
    app.run(debug=True)